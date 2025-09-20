import re
import openpyxl
from openpyxl.utils import get_column_letter

def insert_txt_to_excel(txt_file, excel_file):
    try:
        # 打开Excel文件
        workbook = openpyxl.load_workbook(excel_file)
        # 选择第一个工作表
        sheet = workbook.active
        
        # 读取txt文件内容
        with open(txt_file, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        # 处理每一行
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            # 使用正则表达式提取开头的数字和后面的内容
            match = re.match(r'^(\d+)\.(.*)$', line)
            if match:
                # 提取数字和内容
                num = int(match.group(1))
                content = match.group(2).strip()
                
                # 计算要插入的行号（数字+1）
                row = num + 1
                # 第二列（B列），这里修改列数
                col = 3
                
                # 插入内容
                sheet.cell(row=row, column=col).value = content
                print(f"已插入: 行 {row}, 列 {get_column_letter(col)} - {content}")
            else:
                print(f"跳过无效格式的行: {line}")
        
        # 保存Excel文件
        workbook.save(excel_file)
        print(f"操作完成，已保存到 {excel_file}")
        
    except Exception as e:
        print(f"发生错误: {str(e)}")

if __name__ == "__main__":
    # 输入文件路径
    txt_path = './100things-EN.txt'
    excel_path = './100things.xlsx'
    
    # 调用函数执行操作
    insert_txt_to_excel(txt_path, excel_path)
    