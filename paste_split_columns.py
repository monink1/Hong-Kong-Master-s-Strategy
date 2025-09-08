import openpyxl
import re

def import_txt_to_excel(txt_file, excel_file):
    try:
        # 打开Excel文件，如果文件不存在则创建新的
        try:
            workbook = openpyxl.load_workbook(excel_file)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
        
        # 选择第一个工作表
        sheet = workbook.active
        
        # 读取txt文件内容
        with open(txt_file, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        # 处理每一行，导入到对应的行
        for row_idx, line in enumerate(lines, start=1):  # 从第1行开始
            line = line.strip()
            if not line:
                continue
            
            # 按竖线分割内容，同时处理可能的空格
            parts = [part.strip() for part in re.split(r'\s*\|\s*', line) if part.strip()]
            
            # 将分割后的内容放入不同的列
            for col_idx, part in enumerate(parts, start=1):  # 从第1列开始
                sheet.cell(row=row_idx, column=col_idx).value = part
            
            print(f"已导入行 {row_idx}: {parts}")
        
        # 保存Excel文件
        workbook.save(excel_file)
        print(f"操作完成，已保存到 {excel_file}")
        
    except Exception as e:
        print(f"发生错误: {str(e)}")

if __name__ == "__main__":
    # 输入文件路径
    txt_path = './app.txt'
    excel_path = './app.xlsx'
    
    # 调用函数执行操作
    import_txt_to_excel(txt_path, excel_path)
    